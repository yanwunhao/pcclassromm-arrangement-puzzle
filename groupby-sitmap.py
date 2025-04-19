import csv
import random


def convert_to_2D_array_transposed(input_array, rows=12, cols=2):
    """
    Convert a 1D array into a transposed 2D array with specified rows and columns.
    Handles cases where input array doesn't have enough elements.
    Transposes the result (swaps rows and columns).

    Parameters:
    input_array (list): A 1D array or list of elements
    rows (int): Number of rows in the intermediate array (default: 12)
    cols (int): Number of columns in the intermediate array (default: 2)

    Returns:
    list: A transposed 2D list where the original rows become columns and vice versa
    """
    # Initialize the 2D array with None values
    intermediate = [[None for _ in range(cols)] for _ in range(rows)]

    # Calculate the total cells in the 2D array
    total_cells = rows * cols

    # Fill the 2D array with available elements
    for i in range(min(len(input_array), total_cells)):
        row_idx = i // cols
        col_idx = i % cols
        intermediate[row_idx][col_idx] = input_array[i]

    # Transpose the result (swap rows and columns)
    transposed = [[None for _ in range(rows)] for _ in range(cols)]
    for i in range(rows):
        for j in range(cols):
            transposed[j][i] = intermediate[i][j]

    return transposed


def csv_to_list(file_path):
    result_list = []

    with open(file_path, "r", newline="") as file:
        csv_reader = csv.reader(file)
        for row in csv_reader:
            if row:  # Check if row is not empty
                result_list.append(row[0])

    return result_list


def pull_24people(input_list):
    if len(input_list) < 24:
        pulled_elements = random.sample(input_list, len(input_list))
    else:
        # Pull 24 random elements from the input list
        pulled_elements = random.sample(input_list, 24)

    # Remove the pulled elements from the original list
    for element in pulled_elements:
        input_list.remove(element)

    return pulled_elements


if __name__ == "__main__":
    class1_list = csv_to_list("./data/class1.csv")
    class2_list = csv_to_list("./data/class2.csv")

    new_class1_list = []
    new_class2_list = []

    for item in class1_list:
        item = item + "_class1"
        new_class1_list.append(item)
    class1_list = new_class1_list
    for item in class2_list:
        item = item + "_class2"
        new_class2_list.append(item)
    class2_list = new_class2_list

    sitmap = []
    i = 0
    while class1_list or class2_list:
        group_name = "Class 1" if i % 2 == 0 else "Class 2"
        pull_group = pull_24people(class1_list if i % 2 == 0 else class2_list)

        # Convert the pulled group to 12x2 format
        group_12x2 = convert_to_2D_array_transposed(pull_group)

        # Print group header
        print(f"\n{'-'*50}")
        print(f"Group {i+1} - {group_name}")
        print(f"{'-'*50}")

        # Print the 2D array in a readable format
        for row_idx, row in enumerate(group_12x2):
            formatted_row = []
            for item in row:
                if item is None:
                    formatted_row.append("Empty")
                else:
                    # Extract just the name without the class suffix for cleaner display
                    name = item.split("_")[0] if item else "Empty"
                    formatted_row.append(name)

            print(f"Row {row_idx+1}: {formatted_row}")

        # Add group to sitmap for further processing if needed
        sitmap.append(group_12x2)
        sitmap.append([""])
        i += 1

    # If you want to export to Excel
    try:
        import openpyxl

        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Seating Map"

        row_offset = 1
        for group_idx, group in enumerate(sitmap):
            if group == [""]:
                # This is a separator row
                row_offset += 1
                continue

            # Add group header
            group_name = "Class 1" if group_idx % 2 == 0 else "Class 2"
            sheet.cell(row=row_offset, column=1).value = (
                f"Group {group_idx//2+1} - {group_name}"
            )
            row_offset += 1

            # Add the data
            for row_idx, row in enumerate(group):
                for col_idx, value in enumerate(row):
                    name = value.split("_")[0] if value else "Empty"
                    sheet.cell(row=row_offset + row_idx, column=col_idx + 1).value = (
                        name
                    )

            row_offset += len(group) + 1  # Add space between groups

        workbook.save("seating_map.xlsx")
        print("\nSeating map has been exported to seating_map.xlsx")
    except ImportError:
        print("\nNote: Install openpyxl package to export to Excel")
